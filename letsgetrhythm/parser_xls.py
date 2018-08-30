#-*- coding: utf-8 -*-
import operator
import datetime
import time
import re

from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import render_to_response, redirect, get_object_or_404
from django.core.urlresolvers import reverse
from django.conf import settings
from django.views.decorators.cache import never_cache
from django.template.context import RequestContext

from openpyxl import load_workbook
from unidecode import unidecode

from base.models import *
from user_registration.func import *
from user_registration.ajax import create_by_email
from kinoinfo_folder.func import low, del_separator
from organizations.parser import org_build_create


def parse_address(data):
    num, name, atype = ('','','')
    
    if data:
        num = re.findall(ur'^\d+\-?\d*\w?\s', data)
        num = num[0].strip() if num else ''
        
        data = re.sub(ur'^\d+\-?\d*\w?\s', '', data)
        
        types = {
            'st': 70, 
            'street': 70, 
            'rd': 71, 
            'road': 71, 
            'rds': 71, 
            'ave': 72, 
            'avenue': 72, 
            'hwy': 73, 
            'dr': 74, 
            'crt': 75, 
            'pde': 76
        }
        
        low_data = low(data)

        for k, v in types.iteritems():
            tmp = '%s.' % k
            if k in low_data or tmp in low_data:
                atype = v
        
        if not atype:
            atype = 70
        
        name = re.sub(r'(Avenue|avenue|Ave|ave|Street|street|Road|road|Rds|Hwy|hwy|Crt|crt|Pde|pde|St|st|Rd|rd|Dr)\.?$', '', data)

    return num, name.strip(), atype


@only_superuser_or_admin
@never_cache
def parse_xls(request):
    
    current_site = request.current_site
    
    areas = {}
    for i in Area.objects.all():
        areas[i.slug] = i
    
    orgs = Organization.objects.filter(domain=current_site)
    orgs_dict = {}
    for i in orgs:
        orgs_dict[i.slug] = i
    
    org_phones = OrganizationPhones.objects.filter(organization__domain=current_site)
    phones_objs = {}
    for i in org_phones:
        phones_objs[i.phone] = i
    
    org_tags = OrganizationTags.objects.filter(organization__domain=current_site)
    tags_objs = {}
    for i in org_tags:
        tags_objs[i.name] = i
    
    org_streets = Street.objects.all()
    org_streets_dict = {}
    for i in org_streets:
        org_streets_dict[i.slug.encode('utf-8')] = i
    
    city = City.objects.get(name__name='Melbourne', name__status=1)
    
    
    wb2 = load_workbook('%s/letsget_clients_bank2.xlsx' % settings.API_EX_PATH, use_iterators=True)
    p = wb2.get_active_sheet()
    
    call_results = {
        'FFC00000': False,
        'FFFF0000': False,
        'FFFFFF00': True,
    }
    
    for ind, i in enumerate(p.iter_rows()):
        if ind > 0:
            if i[0].row > 373:
                break
                
            call_result = call_results.get(i[1].style.fill.fgColor.value)
            area = i[3].value.strip().encode('utf-8') if i[3].value else None
            name = i[1].value.strip().encode('utf-8') if i[1].value else area
            code = i[0].value.strip() if i[0].value else name.decode('utf-8')
            slug = low(del_separator(name))
            address = i[2].value.strip() if i[2].value else None
            phone = str(i[4].value).strip().replace(' ','')[:8] if i[4].value else None
            dtime = str(i[6].value).strip().replace(' 00:00:00','') if i[6].value else ''

            try:
                email = str(i[7].value).strip() if i[7].value else ''
            except UnicodeEncodeError:
                email = str(i[7].value.encode('utf-8')).strip() if i[7].value else ''
                
            try:
                note_1 = str(i[8].value).strip() if i[8].value else ''
            except UnicodeEncodeError:
                note_1 = str(i[8].value.encode('utf-8')).strip() if i[8].value else ''
                
            note_2 = str(i[9].value).strip() if i[9].value else ''

            try:
                contact = str(i[5].value).strip() if i[5].value else None
            except UnicodeEncodeError:
                contact = i[5].value.encode('utf-8')
            
            org_obj = orgs_dict.get(slug.decode('utf-8'))
            

            if not '@' in email:
                if not note_1 and email:
                    note_1 = email
                elif not note_2 and email:
                    note_2 = email
                email = None

            extra = '%s;%s;%s;%s' % (dtime, note_1, note_2, call_result)

            house, addr_name, addr_type = parse_address(address)

            if addr_name:
                addr_slug = low(del_separator(addr_name.encode('utf-8')))
                street_obj = org_streets_dict.get(addr_slug)
                if not street_obj:
                    area_obj = None
                    if area:
                        area_slug = low(del_separator(area))
                        area_obj = areas.get(area_slug.decode('utf-8'))
                        if not area_obj:
                            area_obj = Area.objects.create(name=area, slug=area_slug)
                            areas[area_slug.decode('utf-8')] = area_obj
                        
                    street_obj = Street.objects.create(name=addr_name, slug=addr_slug, type=addr_type, area=area_obj)
                    org_streets_dict[addr_slug] = street_obj
            else:
                street_obj = None
                house = None
            
            building_obj = org_build_create(house, city, street_obj)
            
            phone_obj = None
            if phone:
                phone_obj = phones_objs.get(phone.decode('utf-8'))
                if not phone_obj:
                    phone_obj = OrganizationPhones.objects.create(phone=phone)
                    phones_objs[phone.decode('utf-8')] = phone_obj
                
            tag_obj = tags_objs.get(code)
            if not tag_obj:
                tag_obj = OrganizationTags.objects.create(name=code)
                tags_objs[code] = tag_obj
            
            
            if not org_obj:
                org_obj = Organization.objects.create(
                    name = name,
                    slug = slug,
                    source_id = code,
                    email = email,
                    domain = current_site,
                    extra = extra,
                )
            
            if phone_obj and phone_obj not in org_obj.phones.all():
                org_obj.phones.add(phone_obj)

            if building_obj:
                build_tmp = org_obj.buildings.all()
                if build_tmp:
                    if build_tmp[0].number is None and building_obj.number:
                        org_obj.buildings.remove(build_tmp[0])
                        org_obj.buildings.add(building_obj)
                else:
                    org_obj.buildings.add(building_obj)

            
            if tag_obj:
                Organization_Tags.objects.get_or_create(
                    organization = org_obj,
                    organizationtags = tag_obj,
                    defaults = {
                        'organization': org_obj,
                        'organizationtags': tag_obj
                })

            result = unidecode(name.decode('utf-8'))
            result = re.findall(ur'[a-z0-9]+', low(result))
            result = '-'.join(result) if result else ''
            org_obj.uni_slug = '%s-%s' % (result, org_obj.id)
            org_obj.save()


            staff = org_obj.staff.all()
            if staff:
                profile = staff[0]
                if email:
                    if not profile.accounts.all():
                        from user_registration.func import md5_string_generate
                        val_code = md5_string_generate(email)
                        
                        acc, created = Accounts.objects.get_or_create(
                            login = email,
                            defaults = {
                            'login': email,
                            'validation_code': val_code,
                            'email': email,
                        })
                        profile.accounts.add(acc)
                
                editors_all = org_obj.editors.all()
                if not editors_all:
                    org_obj.editors.add(profile)
            else:
                profile, valid_code = create_by_email(email)
                org_obj.staff.add(profile)
                org_obj.editors.add(profile)
                
            profile.phone = phone
            profile.save()

            if contact:
                p_names = profile.person
                if not p_names.name.all():
                    person_name, created_name = NamePerson.objects.get_or_create(
                        status = 1,
                        language_id = 2,
                        name = contact,
                        defaults = {
                            'status': 1,
                            'language_id': 2,
                            'name': contact,
                        })

                    p_names.name.add(person_name)
            

            LetsGetClients.objects.get_or_create(
                site = current_site,
                organization = org_obj,
                defaults = {
                    'site': current_site,
                    'organization': org_obj,
                })
            
    return HttpResponse(str())


